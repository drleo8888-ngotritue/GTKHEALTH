#!/usr/bin/env python3
"""
GVC Medical Import Tool
Input  : File Excel của từng trạm (STT|Ngày|MãNV|Họtên|Chẩnđoán|Nhómbệnh|Giờvào|Giờra | Thuốc...)
         Header 8 dòng: Tên thuốc / Đơn vị / Hạn SD / Tồn đầu kỳ / Nhập / Xuất / SD / Tồn cuối
Output : File 2 — Báo cáo mặt bệnh (bảng + line chart)
         File 3 — Báo cáo sử dụng thuốc (theo khu)
Yêu cầu: pip install openpyxl
"""

import os, sys, re, sqlite3, unicodedata, difflib
from datetime import datetime, date, timedelta
from pathlib import Path
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.chart.series import DataPoint
except ImportError:
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], check=True)
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, Reference

# ─── Config ──────────────────────────────────────────────────────────────────

DB_PATH = Path(os.path.dirname(os.path.abspath(__file__))) / "gvc_data.db"

# Nhóm bệnh cố định — đúng thứ tự hiển thị trên báo cáo
DISEASE_GROUPS = [
    "HÔ HẤP", "THẦN KINH", "TIÊU HÓA", "CƠ XƯƠNG KHỚP", "NỘI KHOA",
    "RĂNG HÀM MẶT", "TAI MŨI HỌNG", "DA LIỄU", "NGOẠI KHOA", "MẮT",
    "PHỤ KHOA", "SẢN KHOA", "TIẾT NIỆU", "TRUYỀN NHIỄM", "BỆNH KHÁC",
]

# Keywords nhận diện cột bệnh nhân
PATIENT_COL_KW = {
    "stt":          ["stt", "số thứ tự", "#", "序号"],
    "date":         ["ngày", "ngay", "date", "日期", "ngày tháng"],
    "patient_id":   ["mã nv", "manv", "工号", "employee", "ma nv", "mã nhân viên"],
    "patient_name": ["họ tên", "ho ten", "姓名"],
    "diagnosis":    ["chẩn đoán", "chan doan", "diagnosis", "triệu chứng", "trieu chung"],
    "disease_group":["nhóm bệnh", "nhom benh", "phân loại", "科别"],
    "time_in":      ["giờ vào", "gio vao", "进入时间", "giờ đến", "check in"],
    "time_out":     ["giờ ra", "gio ra", "离开时间", "giờ về", "check out"],
    "note":         ["ghi chú", "ghi chu", "chú thích", "备注", "ghi chú/chuyển viện"],
}

PATIENT_FIELD_LABELS = {
    "stt": "STT",
    "date": "Ngày tháng",
    "patient_id": "Mã NV",
    "patient_name": "Họ tên",
    "diagnosis": "Chẩn đoán",
    "disease_group": "Nhóm bệnh",
    "time_in": "Giờ vào",
    "time_out": "Giờ ra",
    "note": "Ghi chú",
    "_ignore": "(Bỏ qua)",
}

MED_FIELD_LABELS = {
    "name":      "Tên thuốc",
    "unit":      "Đơn vị",
    "expiry":    "Hạn sử dụng",
    "beg_stock": "Tồn đầu kỳ",
    "received":  "Nhập trong kỳ",
    "issued":    "Xuất trong kỳ",
    "used":      "Sử dụng trong kỳ",
    "end_stock": "Tồn cuối kỳ",
    "_ignore":   "(Bỏ qua)",
}

# Keywords nhận diện dòng header thuốc
MED_HEADER_KW = {
    "name":       ["tên thuốc", "ten thuoc", "药名", "thuốc"],
    "unit":       ["đơn vị", "don vi", "单元", "单位"],
    "expiry":     ["hạn", "han sd", "hạn sử dụng", "expiry", "hsd", "date"],
    "beg_stock":  ["tồn đầu", "ton dau", "库存", "đầu kỳ", "tồn đk"],
    "received":   ["nhập", "nhap", "入库", "实际入库", "nhận"],
    "issued":     ["xuất", "xuat", "xuất trong", "出库"],
    "used":       ["sử dụng", "su dung", "使用", "sd trong"],
    "end_stock":  ["tồn cuối", "ton cuoi", "cuối kỳ", "tồn ck"],
}

C_GREEN  = "#1F6B3A"
C_BLUE   = "#1D4ED8"
C_AMBER  = "#D97706"
C_RED    = "#DC2626"
C_GRAY   = "#6B7280"
C_BG     = "#F9FAFB"

# ─── Helpers ─────────────────────────────────────────────────────────────────

def norm(s: str) -> str:
    nfd = unicodedata.normalize("NFD", str(s or "").lower())
    return "".join(c for c in nfd if not unicodedata.combining(c)).strip()


def kw_match(text: str, keywords: list) -> bool:
    n = norm(text)
    return any(norm(k) in n or n == norm(k) for k in keywords)


def parse_date(v) -> date | None:
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    if isinstance(v, (int, float)):
        n = int(v)
        if 36526 <= n <= 54787:
            try: return (datetime(1899, 12, 30) + timedelta(days=n)).date()
            except: return None
        return None
    s = str(v).strip()
    for pat in [r"^(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})$",
                r"^(\d{4})[/\-](\d{1,2})[/\-](\d{1,2})$"]:
        m = re.match(pat, s)
        if m:
            g = [int(x) for x in m.groups()]
            yr, mo, dy = (g[2], g[1], g[0]) if g[2] > 100 else (g[0], g[1], g[2])
            if 2000 <= yr <= 2040 and 1 <= mo <= 12 and 1 <= dy <= 31:
                try: return date(yr, mo, dy)
                except: return None
    return None


def skip_row(pid) -> bool:
    s = str(pid or "").strip()
    return not s or s.startswith("#") or s.upper() in ("#N/A", "N/A", "NA")


def thin_border(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

# ─── Database ─────────────────────────────────────────────────────────────────

SCHEMA = """
CREATE TABLE IF NOT EXISTS medicines (
    id   INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT UNIQUE NOT NULL COLLATE NOCASE,
    unit TEXT NOT NULL DEFAULT 'Viên'
);
CREATE TABLE IF NOT EXISTS medicine_aliases (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    medicine_id INTEGER NOT NULL REFERENCES medicines(id),
    alias       TEXT NOT NULL COLLATE NOCASE,
    UNIQUE(alias)
);
CREATE TABLE IF NOT EXISTS patients (
    id   TEXT PRIMARY KEY,
    name TEXT NOT NULL DEFAULT ''
);
CREATE TABLE IF NOT EXISTS encounters (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    patient_id   TEXT NOT NULL,
    patient_name TEXT NOT NULL DEFAULT '',
    enc_date     TEXT NOT NULL,
    time_in      TEXT NOT NULL DEFAULT '',
    time_out     TEXT NOT NULL DEFAULT '',
    diagnosis    TEXT NOT NULL DEFAULT '',
    disease_group TEXT NOT NULL DEFAULT '',
    status       TEXT NOT NULL DEFAULT 'Về làm việc',
    station      TEXT NOT NULL DEFAULT '',
    source_file  TEXT NOT NULL DEFAULT ''
);
CREATE TABLE IF NOT EXISTS prescriptions (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    encounter_id INTEGER NOT NULL REFERENCES encounters(id),
    medicine_id  INTEGER NOT NULL REFERENCES medicines(id),
    quantity     REAL NOT NULL,
    station      TEXT NOT NULL DEFAULT ''
);
CREATE TABLE IF NOT EXISTS station_medicine_meta (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    station     TEXT NOT NULL,
    period      TEXT NOT NULL,
    medicine_id INTEGER NOT NULL REFERENCES medicines(id),
    expiry_date TEXT NOT NULL DEFAULT '',
    beg_stock   REAL NOT NULL DEFAULT 0,
    received    REAL NOT NULL DEFAULT 0,
    issued      REAL NOT NULL DEFAULT 0,
    UNIQUE(station, period, medicine_id)
);
CREATE TABLE IF NOT EXISTS imported_files (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    filename    TEXT NOT NULL,
    station     TEXT NOT NULL,
    period      TEXT NOT NULL,
    imported_at TEXT NOT NULL,
    UNIQUE(filename, station, period)
);
"""


def get_db() -> sqlite3.Connection:
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    conn.executescript(SCHEMA)
    # Migration: thêm cột issued nếu DB cũ chưa có
    try:
        conn.execute("ALTER TABLE station_medicine_meta ADD COLUMN issued REAL NOT NULL DEFAULT 0")
        conn.commit()
    except Exception:
        pass
    conn.commit()
    return conn


def db_stats(conn) -> dict:
    return {
        "encounters": conn.execute("SELECT COUNT(*) FROM encounters").fetchone()[0],
        "medicines":  conn.execute("SELECT COUNT(*) FROM medicines").fetchone()[0],
        "stations":   conn.execute("SELECT COUNT(DISTINCT station) FROM encounters").fetchone()[0],
    }

# ─── Medicine lookup ──────────────────────────────────────────────────────────

def find_med(conn, name: str) -> int | None:
    r = conn.execute("SELECT id FROM medicines WHERE name=?", (name.strip(),)).fetchone()
    if r: return r["id"]
    r = conn.execute("SELECT medicine_id FROM medicine_aliases WHERE alias=?", (name.strip(),)).fetchone()
    return r["medicine_id"] if r else None


def suggest_med(conn, name: str, cutoff=0.55) -> list:
    n_in = norm(name)
    rows = conn.execute("SELECT id, name FROM medicines").fetchall()
    out = []
    for r in rows:
        nm = norm(r["name"])
        score = difflib.SequenceMatcher(None, n_in, nm).ratio()
        if n_in in nm or nm in n_in: score = max(score, 0.80)
        if score >= cutoff: out.append((r["id"], r["name"], score))
    return sorted(out, key=lambda x: -x[2])


def upsert_med(conn, name: str, unit: str = "Viên") -> int:
    mid = find_med(conn, name)
    if mid: return mid
    conn.execute("INSERT INTO medicines (name, unit) VALUES (?,?)", (name.strip(), unit or "Viên"))
    conn.commit()
    return conn.execute("SELECT id FROM medicines WHERE name=?", (name.strip(),)).fetchone()["id"]


def add_alias(conn, mid: int, alias: str):
    try:
        conn.execute("INSERT OR IGNORE INTO medicine_aliases (medicine_id, alias) VALUES (?,?)",
                     (mid, alias.strip()))
        conn.commit()
    except: pass

# ─── Excel Parser ─────────────────────────────────────────────────────────────

def _load_excel_rows(filepath: str, period: str) -> tuple:
    """Load workbook và trả về (rows, error)."""
    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception as e:
        return [], str(e)
    ws = wb.active
    for sn in wb.sheetnames:
        sn_n = norm(sn)
        if norm(period) in sn_n or any(norm(p) in sn_n for p in period.split(".")):
            ws = wb[sn]; break
    merged_vals: dict[tuple, object] = {}
    for mr in ws.merged_cells.ranges:
        master = ws.cell(mr.min_row, mr.min_col).value
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                merged_vals[(r, c)] = master
    rows = []
    for ws_row in ws.iter_rows(values_only=False):
        rows.append(tuple(
            merged_vals.get((cell.row, cell.column), cell.value)
            for cell in ws_row
        ))
    return rows, None


def _detect_structure(rows: list) -> dict:
    """
    Auto-detect patient columns and medicine header rows from raw Excel rows.
    Returns struct dict used by _parse_from_struct().
    """
    data_start_row = -1
    col_map: dict[str, int] = {}
    header_row_idx = 0

    for ri, row in enumerate(rows):
        cells = [str(c or "").strip() for c in row]
        hits, tmp = 0, {}
        for ci, cell in enumerate(cells):
            for field, kws in PATIENT_COL_KW.items():
                if field not in tmp and kw_match(cell, kws):
                    tmp[field] = ci; hits += 1
        if hits >= 3 and "patient_id" in tmp:
            col_map = tmp; data_start_row = ri + 1; header_row_idx = ri; break

    if data_start_row == -1:
        for ri, row in enumerate(rows):
            for c in row:
                if c and not skip_row(c) and parse_date(c) is not None:
                    data_start_row = ri; break
            if data_start_row >= 0: break

    if data_start_row < 0:
        return {"error": "Không tìm thấy dữ liệu bệnh nhân",
                "col_map": {}, "med_header_row_idx": {},
                "label_col": 8, "med_start_col": 9,
                "actual_data_start": 0, "header_row_idx": 0}

    last_patient_col = max(col_map.values()) if col_map else 7
    label_col     = last_patient_col + 1
    med_start_col = last_patient_col + 2

    actual_data_start = data_start_row
    pid_ci  = col_map.get("patient_id")
    date_ci = col_map.get("date")
    if pid_ci is not None:
        for ri in range(data_start_row, min(data_start_row + 20, len(rows))):
            r    = rows[ri]
            pval = str(r[pid_ci] if pid_ci < len(r) else "").strip()
            dval = r[date_ci] if (date_ci is not None and date_ci < len(r)) else None
            if pval and not skip_row(pval) and parse_date(dval) is not None:
                actual_data_start = ri; break

    med_header_row_idx: dict[str, int] = {}
    for ri in range(actual_data_start):
        row   = rows[ri]
        cells = [str(c or "").strip() for c in row]
        row_label = cells[label_col] if label_col < len(cells) else ""
        if not row_label:
            for ci in range(min(label_col, len(cells))):
                if cells[ci]: row_label = cells[ci]; break
        for field, kws in MED_HEADER_KW.items():
            if field not in med_header_row_idx and kw_match(row_label, kws):
                med_header_row_idx[field] = ri; break

    return {
        "col_map":             col_map,
        "med_header_row_idx":  med_header_row_idx,
        "label_col":           label_col,
        "med_start_col":       med_start_col,
        "actual_data_start":   actual_data_start,
        "header_row_idx":      header_row_idx,
        "error":               None,
    }


def _parse_from_struct(rows: list, struct: dict, station: str, period: str) -> dict:
    """Parse encounters và medicines từ raw rows + confirmed struct."""
    col_map            = struct["col_map"]
    med_header_row_idx = struct["med_header_row_idx"]
    label_col          = struct["label_col"]
    med_start_col      = struct["med_start_col"]
    actual_data_start  = struct["actual_data_start"]

    # Build med_header từ confirmed row indices
    med_header: dict[str, dict] = {}
    for field, ri in med_header_row_idx.items():
        if ri < len(rows):
            row = rows[ri]
            med_header[field] = {
                ci: row[ci]
                for ci in range(med_start_col, len(row))
                if row[ci] is not None and str(row[ci]).strip()
            }

    med_names_row = med_header.get("name", {})
    if not med_names_row:
        hrow = rows[struct.get("header_row_idx", 0)] if rows else []
        for ci in range(med_start_col, len(hrow)):
            v = str(hrow[ci] or "").strip()
            if v: med_names_row[ci] = v

    medicines = []
    med_col_map: dict[int, str] = {}
    _header_exact = {norm(k) for kws in MED_HEADER_KW.values() for k in kws}

    for ci, med_name_raw in med_names_row.items():
        med_name = str(med_name_raw).strip()
        if not med_name or len(med_name) < 2: continue
        try:
            float(med_name.replace(",", ".").replace(" ", "")); continue
        except ValueError:
            pass
        _no_cjk = "".join(c for c in med_name if ord(c) < 0x3000).strip()
        if norm(_no_cjk) in _header_exact or norm(med_name) in _header_exact: continue
        med_col_map[ci] = med_name
        medicines.append({
            "name":        med_name,
            "unit":        str(med_header.get("unit",      {}).get(ci, "Viên")).strip() or "Viên",
            "expiry":      str(med_header.get("expiry",    {}).get(ci, "")).strip(),
            "beg_stock":   _to_float(med_header.get("beg_stock",  {}).get(ci, 0)),
            "received":    _to_float(med_header.get("received",   {}).get(ci, 0)),
            "issued":      _to_float(med_header.get("issued",     {}).get(ci, 0)),
            "used_header": _to_float(med_header.get("used",       {}).get(ci, 0)),
        })

    encounters = []
    for row in rows[actual_data_start:]:
        if not any(row): continue

        def gcol(field):
            ci = col_map.get(field)
            if ci is None or ci >= len(row): return ""
            return row[ci]

        pid = str(gcol("patient_id") or "").strip()
        if skip_row(pid): continue

        enc_date = parse_date(gcol("date")) or date.today()

        meds_used: dict[str, float] = {}
        for ci, med_name in med_col_map.items():
            if ci < len(row) and row[ci]:
                qty = _to_float(row[ci])
                if qty > 0:
                    meds_used[med_name] = qty

        status_raw = ""
        if "status" in col_map:
            status_raw = str(gcol("status") or "").strip().lower()
        if not status_raw and "note" in col_map:
            status_raw = str(gcol("note") or "").strip().lower()

        disease_group = str(gcol("disease_group") or "").strip().upper()

        encounters.append({
            "patient_id":    pid,
            "patient_name":  str(gcol("patient_name") or "").strip(),
            "date":          enc_date,
            "diagnosis":     str(gcol("diagnosis") or "").strip(),
            "disease_group": disease_group,
            "time_in":       _format_time(gcol("time_in")),
            "time_out":      _format_time(gcol("time_out")),
            "status":        "Chuyển viện" if status_raw == "x" else "Về làm việc",
            "medicines":     meds_used,
        })

    return {
        "encounters":        encounters,
        "medicines":         medicines,
        "med_col_map":       med_col_map,
        "med_start_col":     med_start_col,
        "actual_data_start": actual_data_start,
        "error":             None,
    }


def parse_station_excel(filepath: str, station: str, period: str,
                        struct: dict = None) -> dict:
    """Đọc 1 file Excel của 1 trạm. struct=None → auto-detect."""
    rows, err = _load_excel_rows(filepath, period)
    if err:
        return {"encounters": [], "medicines": [], "med_col_map": {}, "error": err}
    if not rows:
        return {"encounters": [], "medicines": [], "med_col_map": {}, "error": "Sheet rỗng"}
    if struct is None:
        struct = _detect_structure(rows)
    if struct.get("error"):
        return {"encounters": [], "medicines": [], "med_col_map": {}, "error": struct["error"]}
    return _parse_from_struct(rows, struct, station, period)


def _to_float(v) -> float:
    try: return float(str(v or 0).replace(",", ".").replace(" ", ""))
    except: return 0.0


def _format_time(v) -> str:
    if v is None: return ""
    if isinstance(v, (datetime,)):
        return v.strftime("%H:%M")
    if isinstance(v, (int, float)):
        frac = v % 1
        total_min = round(frac * 1440)
        return f"{total_min//60:02d}:{total_min%60:02d}"
    s = str(v).strip()
    m = re.match(r"(\d{1,2}):(\d{2})", s)
    return f"{int(m[1]):02d}:{m[2]}" if m else s


def _normalize_disease_group(raw: str) -> str:
    """Map tên nhóm bệnh thực tế về danh sách chuẩn."""
    n = norm(raw)
    for g in DISEASE_GROUPS:
        if norm(g) in n or n in norm(g) or n == norm(g):
            return g
    # Một số alias thường gặp
    aliases = {
        "ho hap": "HÔ HẤP", "than kinh": "THẦN KINH", "tieu hoa": "TIÊU HÓA",
        "co xuong": "CƠ XƯƠNG KHỚP", "noi khoa": "NỘI KHOA",
        "rang ham mat": "RĂNG HÀM MẶT", "tai mui hong": "TAI MŨI HỌNG",
        "da lieu": "DA LIỄU", "ngoai khoa": "NGOẠI KHOA",
        "phu khoa": "PHỤ KHOA", "san khoa": "SẢN KHOA",
        "tiet nieu": "TIẾT NIỆU", "truyen nhiem": "TRUYỀN NHIỄM",
    }
    for k, v in aliases.items():
        if k in n: return v
    return raw.upper() if raw else "BỆNH KHÁC"

# ─── Report: Tổng hợp bệnh nhân (File 1) ─────────────────────────────────────

def generate_patient_report(output_path: str, period_label: str = ""):
    conn = get_db()
    try:
        rows = conn.execute("""
            SELECT enc_date, patient_id, patient_name, station,
                   diagnosis, disease_group, status
            FROM encounters
            ORDER BY enc_date, station, patient_id
        """).fetchall()
    finally:
        conn.close()

    # Đếm tổng số lần vào y tế theo (ID, tên) — cùng ID nhưng khác tên = người khác
    visit_total: dict[tuple, int] = {}
    for r in rows:
        key = (r["patient_id"], r["patient_name"])
        visit_total[key] = visit_total.get(key, 0) + 1

    wb = Workbook()
    ws = wb.active
    ws.title = "Theo dõi bệnh nhân"

    H_FILL = PatternFill("solid", fgColor="1F6B3A")
    H_FONT = Font(color="FFFFFF", bold=True, name="Arial", size=10)
    ALT    = PatternFill("solid", fgColor="F0FDF4")

    # ── Dòng 1: tiêu đề ──────────────────────────────────────────────────────
    title = f"BẢNG THEO DÕI CBNV CÔNG TY GOERTEK VINA SỬ DỤNG PHÒNG Y TẾ {period_label}".strip()
    ws.merge_cells("A1:J1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13, name="Arial", color="1F6B3A")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].fill = PatternFill("solid", fgColor="DCFCE7")
    ws.row_dimensions[1].height = 30

    # ── Dòng 2: header ───────────────────────────────────────────────────────
    headers = [
        ("STT\n序号",             6),
        ("Ngày tháng\n日期",      14),
        ("Mã NV\n工号",           12),
        ("Họ tên\n姓名",          22),
        ("Bộ phận\n部门",         18),
        ("Triệu chứng ban đầu\n临床表现", 32),
        ("Nhóm bệnh\n疾病组",     16),
        ("Khu",                   8),
        ("Ghi chú",               14),
        ("Số lần\nvào y tế",      10),
    ]
    for ci, (h, w) in enumerate(headers, 1):
        c = ws.cell(2, ci, h)
        c.fill = H_FILL; c.font = H_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border("AAAAAA")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 34

    # ── Dòng 3+: dữ liệu ─────────────────────────────────────────────────────
    CENTER_COLS = {1, 2, 3, 8, 10}
    for idx, r in enumerate(rows, 3):
        fill = ALT if idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        # Ghi chú: chỉ ghi nếu chuyển viện
        note = "Chuyển viện" if r["status"] == "Chuyển viện" else ""
        vals = [
            idx - 2,
            r["enc_date"],
            r["patient_id"],
            r["patient_name"],
            "",              # Bộ phận — không có trong DB
            r["diagnosis"],
            r["disease_group"],
            r["station"],
            note,
            visit_total.get((r["patient_id"], r["patient_name"]), 1),
        ]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(idx, ci, val)
            c.fill = fill
            c.border = thin_border()
            c.alignment = Alignment(
                horizontal="center" if ci in CENTER_COLS else "left",
                vertical="center")

    # ── Dòng tổng ─────────────────────────────────────────────────────────────
    total_ri = len(rows) + 3
    ws.merge_cells(f"A{total_ri}:J{total_ri}")
    ws.cell(total_ri, 1, f"TỔNG CỘNG: {len(rows)} lượt khám")
    ws.cell(total_ri, 1).font = Font(bold=True, name="Arial", color="1F6B3A")
    ws.cell(total_ri, 1).fill = PatternFill("solid", fgColor="DCFCE7")
    ws.cell(total_ri, 1).alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A3"
    wb.save(output_path)


# ─── Template ─────────────────────────────────────────────────────────────────

def generate_template(output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "THÁNG"

    GRN = PatternFill("solid", fgColor="1F6B3A")
    BLU = PatternFill("solid", fgColor="1D4ED8")
    YEL = PatternFill("solid", fgColor="FEF9C3")
    SAM = PatternFill("solid", fgColor="F0FDF4")
    WFT = lambda bold=True: Font(color="FFFFFF", bold=bold, name="Arial", size=9)

    # ── Dòng 1–7: header thuốc (ví dụ) ──────────────────────────────────────
    med_start_col = 9  # Cột I

    # Dòng 1: Tên cột bệnh nhân
    patient_headers = [
        ("STT", 6), ("Ngày tháng", 14), ("Mã NV", 10), ("Họ tên", 22),
        ("Chẩn đoán", 28), ("Nhóm bệnh", 18), ("Giờ vào", 10), ("Giờ ra", 10),
    ]
    for ci, (h, w) in enumerate(patient_headers, 1):
        c = ws.cell(1, ci, h)
        c.fill = GRN; c.font = WFT()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border("FFFFFF")
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Dòng header thuốc theo chiều dọc
    med_row_labels = [
        ("Tên thuốc",        1, GRN),
        ("Đơn vị",           2, GRN),
        ("Hạn sử dụng",      3, GRN),
        ("Tồn đầu kỳ",       4, PatternFill("solid", fgColor="065F46")),
        ("Nhập trong kỳ\n实际入库数量", 5, PatternFill("solid", fgColor="065F46")),
        ("Xuất trong kỳ",    6, PatternFill("solid", fgColor="065F46")),
        ("Sử dụng trong kỳ\n使用", 7, PatternFill("solid", fgColor="92400E")),
        ("Tồn cuối kỳ",      8, PatternFill("solid", fgColor="1E3A8A")),
    ]
    for label, row_offset, fill in med_row_labels:
        c = ws.cell(row_offset, med_start_col - 1, label)
        c.fill = fill; c.font = WFT(); c.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[row_offset].height = 18

    # Ví dụ 2 cột thuốc mẫu
    sample_meds = [
        ("Amoxicillin 500mg", "Viên", "04/04/2027", 1000, 0),
        ("[Tên thuốc 2]",     "Viên", "dd/MM/yyyy", 0,    0),
    ]
    for mi, (name, unit, exp, beg, recv) in enumerate(sample_meds):
        ci = med_start_col + mi
        ws.cell(1, ci, name).fill = BLU; ws.cell(1, ci).font = WFT()
        ws.cell(1, ci).alignment = Alignment(horizontal="center", wrap_text=True)
        ws.cell(2, ci, unit)
        ws.cell(3, ci, exp)
        ws.cell(4, ci, beg)
        ws.cell(5, ci, recv)
        ws.cell(6, ci, 0)   # Xuất trong kỳ — nhập tay
        col_letter = get_column_letter(ci)
        ws.cell(7, ci, f"=SUM({col_letter}9:{col_letter}10000)")   # Sử dụng = SUM data rows
        ws.cell(8, ci, f"={col_letter}4+{col_letter}5-{col_letter}6-{col_letter}7")  # Tồn cuối
        ws.column_dimensions[col_letter].width = 14

    # Làm đẹp tồn/nhập/sd/cuối
    for row_i, color in [(4,"E6FAF0"), (5,"E6FAF0"), (6,"E6FAF0"), (7,"FEF3C7"), (8,"DBEAFE")]:
        for ci in range(med_start_col, med_start_col + 2):
            ws.cell(row_i, ci).fill = PatternFill("solid", fgColor=color)
            ws.cell(row_i, ci).alignment = Alignment(horizontal="center")

    # ── Dòng 8+: dữ liệu mẫu ────────────────────────────────────────────────
    sample_data = [
        ["",  1, "01/01/2026", "100001", "NGUYEN VAN A",   "Đau đầu, sổ mũi", "HÔ HẤP",  "08:00", "08:30", 2, ""],
        ["x", 2, "01/01/2026", "100002", "TRAN THI B",     "Đau bụng cấp",    "TIÊU HÓA","09:00", "09:45", "", 3],
    ]
    for ri, row in enumerate(sample_data, 9):
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, val)
            c.fill = SAM; c.alignment = Alignment(horizontal="center")
            c.border = thin_border()

    # Ghi chú
    notes = [
        "LƯU Ý:",
        "• Cột bệnh nhân (A-H): STT | Ngày | Mã NV | Họ tên | Chẩn đoán | Nhóm bệnh | Giờ vào | Giờ ra",
        "• Cột thuốc: bắt đầu từ cột I trở đi. Header 8 dòng: Tên / ĐV / Hạn SD / Tồn đầu / Nhập / Xuất / SD / Tồn cuối",
        "• Dữ liệu bệnh nhân bắt đầu từ dòng 9",
        "• Cột Trạng thái (nếu có): 'x' = chuyển viện, để trống = về làm việc",
    ]
    for i, note in enumerate(notes, 13):
        ws.cell(i, 1, note).font = Font(color="92400E" if i == 13 else "374151",
                                         bold=(i == 13), name="Arial", size=8)
        ws.merge_cells(f"A{i}:{get_column_letter(med_start_col+3)}{i}")
        ws.cell(i, 1).fill = PatternFill("solid", fgColor="FFFBEB")

    ws.row_dimensions[1].height = 24

    # Đóng băng dòng 8 trở đi
    ws.freeze_panes = "A9"

    wb.save(output_path)

# ─── Report: Báo cáo mặt bệnh (File 2) ───────────────────────────────────────

def generate_disease_report(output_path: str, period_label: str = ""):
    conn = get_db()
    try:
        # Tổng ca theo nhóm bệnh
        rows = conn.execute("""
            SELECT disease_group,
                   COUNT(*) AS total,
                   SUM(status='Chuyển viện') AS xfer
            FROM encounters
            GROUP BY disease_group
        """).fetchall()
    finally:
        conn.close()

    # Gom vào dict
    group_map: dict[str, dict] = {g: {"total": 0, "xfer": 0} for g in DISEASE_GROUPS}
    other_total, other_xfer = 0, 0
    for r in rows:
        g = r["disease_group"]
        if g in group_map:
            group_map[g]["total"] += r["total"]
            group_map[g]["xfer"]  += r["xfer"]
        else:
            other_total += r["total"]
            other_xfer  += r["xfer"]
    group_map["BỆNH KHÁC"]["total"] += other_total
    group_map["BỆNH KHÁC"]["xfer"]  += other_xfer

    grand_total = sum(v["total"] for v in group_map.values())
    grand_xfer  = sum(v["xfer"]  for v in group_map.values())

    wb = Workbook()
    ws = wb.active
    ws.title = "Báo cáo mặt bệnh"

    # ── Tiêu đề ──────────────────────────────────────────────────────────────
    title = f"BÁO CÁO MẶT BỆNH PHÒNG Y TẾ {period_label}".strip()
    ws.merge_cells("A1:E1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13, name="Arial", color="1F6B3A")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    # ── Header bảng ──────────────────────────────────────────────────────────
    H_FILL = PatternFill("solid", fgColor="1F6B3A")
    H_FONT = Font(color="FFFFFF", bold=True, name="Arial", size=10)
    ALT    = PatternFill("solid", fgColor="F0FDF4")

    for ci, txt in enumerate(["LOẠI BỆNH", "Số lượng", "Tỷ lệ", "Ghi chú"], 1):
        c = ws.cell(2, ci, txt)
        c.fill = H_FILL; c.font = H_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border("AAAAAA")
    ws.row_dimensions[2].height = 22

    # ── Dữ liệu ──────────────────────────────────────────────────────────────
    # Tất cả 15 nhóm, kể cả nhóm = 0, sắp xếp giảm dần
    data_groups_sorted = sorted(DISEASE_GROUPS, key=lambda g: -group_map[g]["total"])

    for ri, g in enumerate(data_groups_sorted, 3):
        total = group_map[g]["total"]
        pct   = total / grand_total if grand_total else 0
        fill  = ALT if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

        ws.cell(ri, 1, g).fill = fill
        ws.cell(ri, 2, total).fill = fill
        ws.cell(ri, 3, f"{pct:.1%}").fill = fill
        ws.cell(ri, 4, "").fill = fill
        for ci in range(1, 5):
            ws.cell(ri, ci).border = thin_border()
            ws.cell(ri, ci).alignment = Alignment(horizontal="center" if ci > 1 else "left",
                                                   vertical="center")

    # Dòng TỔNG
    total_row = len(data_groups_sorted) + 3
    ws.cell(total_row, 1, "TỔNG").font = Font(bold=True, name="Arial")
    ws.cell(total_row, 2, grand_total).font = Font(bold=True, name="Arial")
    ws.cell(total_row, 3, "100%").font = Font(bold=True, name="Arial")
    for ci in range(1, 5):
        ws.cell(total_row, ci).fill = PatternFill("solid", fgColor="DCFCE7")
        ws.cell(total_row, ci).border = thin_border()
        ws.cell(total_row, ci).alignment = Alignment(horizontal="center" if ci > 1 else "left")

    # Dòng chuyển viện
    xfer_row = total_row + 1
    ws.cell(xfer_row, 1, "Số ca chuyển viện")
    ws.cell(xfer_row, 2, grand_xfer)
    pct_xfer = grand_xfer / grand_total if grand_total else 0
    ws.cell(xfer_row, 3, f"{pct_xfer:.1%}")
    for ci in range(1, 4):
        ws.cell(xfer_row, ci).fill = PatternFill("solid", fgColor="FEF9C3")
        ws.cell(xfer_row, ci).font = Font(bold=True, color="92400E", name="Arial")
        ws.cell(xfer_row, ci).border = thin_border()
        ws.cell(xfer_row, ci).alignment = Alignment(horizontal="center" if ci > 1 else "left")

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 18

    # ── Combo Chart: cột xám (nền) + đường xanh (trên) ──────────────────────
    if data_groups_sorted and grand_total:
        n_groups = len(data_groups_sorted)

        # Cột E ẩn: lưu phân số 0.275 (= 27.5%), format cell '0.0%'
        pct_col = 5
        ws.cell(2, pct_col, "Tỷ lệ")
        for ri, g in enumerate(data_groups_sorted, 3):
            pct_val = round(group_map[g]["total"] / grand_total, 4)
            cell = ws.cell(ri, pct_col, pct_val)
            cell.number_format = '0.0%'
        ws.column_dimensions["E"].width = 0

        data_ref = Reference(ws, min_col=pct_col, min_row=3, max_row=2 + n_groups)
        cats_ref = Reference(ws, min_col=1,       min_row=3, max_row=2 + n_groups)

        # ── BarChart: cột xám làm nền ─────────────────────────────────────────
        bar = BarChart()
        bar.type     = "col"
        bar.grouping = "clustered"
        bar.title    = title
        bar.y_axis.numFmt = '0.0%'
        bar.height   = 16
        bar.width    = 28
        bar.style    = 10
        bar.add_data(data_ref)
        bar.set_categories(cats_ref)
        try:
            bar.series[0].graphicalProperties.solidFill      = "C0C0C0"
            bar.series[0].graphicalProperties.line.solidFill = "C0C0C0"
        except Exception:
            pass

        # ── LineChart: đường xanh + marker + data label ───────────────────────
        line = LineChart()
        line.add_data(data_ref)
        line.set_categories(cats_ref)
        line.series[0].graphicalProperties.line.solidFill = "22C55E"
        line.series[0].graphicalProperties.line.width     = 25000   # 2.5pt
        line.series[0].marker.symbol = "circle"
        line.series[0].marker.size   = 7
        try:
            line.series[0].marker.graphicalProperties.solidFill      = "22C55E"
            line.series[0].marker.graphicalProperties.line.solidFill = "22C55E"
        except Exception:
            pass

        # Data labels — "27.5%" — thử dùng AxisNumFmt, fallback hiện số thuần
        try:
            from openpyxl.chart.label import DataLabelList
            from openpyxl.chart.axis  import NumFmt as AxisNumFmt
            dLbls = DataLabelList()
            dLbls.showVal       = True
            dLbls.showLegendKey = False
            dLbls.showCatName   = False
            dLbls.showSerName   = False
            dLbls.showPercent   = False
            dLbls.numFmt = AxisNumFmt(formatCode='0.0%', sourceLinked=False)
            line.series[0].dLbls = dLbls
        except Exception:
            try:
                from openpyxl.chart.label import DataLabelList
                dLbls = DataLabelList()
                dLbls.showVal = True
                dLbls.showLegendKey = dLbls.showCatName = dLbls.showSerName = False
                line.series[0].dLbls = dLbls
            except Exception:
                pass

        # Overlay line lên bar → combo chart
        bar += line
        ws.add_chart(bar, "F2")

    wb.save(output_path)

# ─── Report: Báo cáo sử dụng thuốc (File 3) ──────────────────────────────────

def generate_medicine_report(output_path: str, period_label: str = ""):
    conn = get_db()
    try:
        stations = [r[0] for r in conn.execute(
            "SELECT DISTINCT station FROM encounters ORDER BY station"
        ).fetchall()]
        medicines = conn.execute("SELECT id, name, unit FROM medicines ORDER BY name").fetchall()
        # Sử dụng trong kỳ theo trạm
        used_data = conn.execute("""
            SELECT p.station, p.medicine_id, SUM(p.quantity) AS total_used
            FROM prescriptions p GROUP BY p.station, p.medicine_id
        """).fetchall()
        # Meta (tồn đầu, nhập)
        meta_data = conn.execute(
            "SELECT station, medicine_id, expiry_date, beg_stock, received, issued FROM station_medicine_meta"
        ).fetchall()
    finally:
        conn.close()

    # Build lookup
    used_map: dict[tuple, float] = {}
    for r in used_data:
        used_map[(r["station"], r["medicine_id"])] = r["total_used"]

    meta_map: dict[tuple, dict] = {}
    for r in meta_data:
        meta_map[(r["station"], r["medicine_id"])] = {
            "expiry":    r["expiry_date"],
            "beg_stock": r["beg_stock"],
            "received":  r["received"],
            "issued":    r["issued"],
        }

    wb = Workbook()
    ws = wb.active
    ws.title = "Báo cáo thuốc"

    # Màu sắc
    H1  = PatternFill("solid", fgColor="1F6B3A")
    H2  = PatternFill("solid", fgColor="065F46")
    H3  = PatternFill("solid", fgColor="1E3A8A")
    WF  = Font(color="FFFFFF", bold=True, name="Arial", size=9)
    ALT = PatternFill("solid", fgColor="F0FAF4")

    # ── Tiêu đề ──────────────────────────────────────────────────────────────
    n_station_cols = len(stations) * 5  # mỗi trạm: Tồn đầu / Nhập / Xuất / SD / Tồn cuối
    ws.merge_cells(f"A1:{get_column_letter(3 + n_station_cols)}1")
    ws["A1"] = f"BÁO CÁO SỬ DỤNG THUỐC {period_label}".strip()
    ws["A1"].font = Font(bold=True, size=13, name="Arial", color="1F6B3A")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    # ── Dòng 2: fixed headers ─────────────────────────────────────────────────
    for ci, (txt, w) in enumerate([("Tên thuốc", 30), ("Đơn vị", 10)], 1):
        c = ws.cell(2, ci, txt); c.fill = H1; c.font = WF
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border("AAAAAA")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.merge_cells("A2:A3"); ws.merge_cells("B2:B3")

    # ── Dòng 2–3: station headers (mỗi trạm gom 4 cột) ──────────────────────
    col_cursor = 3
    station_col_starts: dict[str, int] = {}
    for st in stations:
        station_col_starts[st] = col_cursor
        end_col = col_cursor + 4
        ws.merge_cells(f"{get_column_letter(col_cursor)}2:{get_column_letter(end_col)}2")
        c = ws.cell(2, col_cursor, f"KHU {st}" if not st.upper().startswith("KHU") else st)
        c.fill = H2; c.font = WF; c.alignment = Alignment(horizontal="center")
        c.border = thin_border("AAAAAA")

        sub_headers = ["Tồn đầu kỳ", "Nhập trong kỳ", "Xuất trong kỳ",
                       "Sử dụng trong kỳ", "Tồn cuối kỳ"]
        sub_fills   = [H2, H2, H2,
                       PatternFill("solid", fgColor="92400E"), H3]
        for i, (sh, sf) in enumerate(zip(sub_headers, sub_fills)):
            c2 = ws.cell(3, col_cursor + i, sh)
            c2.fill = sf; c2.font = WF
            c2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c2.border = thin_border("AAAAAA")
            ws.column_dimensions[get_column_letter(col_cursor + i)].width = 14
        col_cursor += 5

    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 28

    # ── Dữ liệu thuốc ────────────────────────────────────────────────────────
    for ri, med in enumerate(medicines, 4):
        fill = ALT if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

        ws.cell(ri, 1, med["name"]).fill = fill
        ws.cell(ri, 2, med["unit"]).fill = fill
        for ci in range(1, 3):
            ws.cell(ri, ci).border = thin_border()
            ws.cell(ri, ci).alignment = Alignment(vertical="center",
                                                   horizontal="left" if ci == 1 else "center")

        for st in stations:
            sc   = station_col_starts[st]
            meta = meta_map.get((st, med["id"]),
                                {"beg_stock": 0, "received": 0, "issued": 0, "expiry": ""})
            used = used_map.get((st, med["id"]), 0.0)
            beg  = meta["beg_stock"]
            recv = meta["received"]
            issd = meta["issued"]
            end  = beg + recv - issd - used

            vals      = [beg,  recv, issd, used,                                        end]
            val_fills = [fill, fill, fill, PatternFill("solid", fgColor="FEF3C7"), fill]
            for i, (val, vf) in enumerate(zip(vals, val_fills)):
                c = ws.cell(ri, sc + i, val if val else "")
                c.fill = vf if val else fill
                c.border = thin_border()
                c.alignment = Alignment(horizontal="center", vertical="center")

    # Dòng tổng cộng
    total_ri = len(medicines) + 4
    ws.cell(total_ri, 1, "TỔNG CỘNG").font = Font(bold=True, name="Arial")
    ws.cell(total_ri, 1).fill = PatternFill("solid", fgColor="DCFCE7")
    ws.cell(total_ri, 2, "").fill = PatternFill("solid", fgColor="DCFCE7")
    for st in stations:
        sc = station_col_starts[st]
        for i in range(5):
            col_letter = get_column_letter(sc + i)
            ws.cell(total_ri, sc + i, f"=SUM({col_letter}4:{col_letter}{total_ri-1})")
            ws.cell(total_ri, sc + i).fill = PatternFill("solid", fgColor="DCFCE7")
            ws.cell(total_ri, sc + i).font = Font(bold=True, name="Arial")
            ws.cell(total_ri, sc + i).border = thin_border()
            ws.cell(total_ri, sc + i).alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A4"
    wb.save(output_path)

# ─── Multi-File Dialog ────────────────────────────────────────────────────────

class MultiFileDialog(tk.Toplevel):
    """Chọn nhiều file, auto-detect tên trạm, cho phép sửa trước khi xác nhận."""

    def __init__(self, parent, filepaths: list):
        super().__init__(parent)
        self.title("Xác nhận trạm và kỳ")
        self.geometry("720x420")
        self.resizable(True, True)
        self.grab_set()
        self.result: list[dict] = []
        self._filepaths = filepaths
        self._station_vars: list[tk.StringVar] = []
        self._build(filepaths)

    @staticmethod
    def _auto_station(stem: str) -> str:
        m = re.search(r"(?:khu[_\s]*)?([A-Za-z]\d+|[A-Za-z]+\s*\d*)", stem, re.IGNORECASE)
        return m.group(1).strip().upper() if m else stem[:8].upper()

    def _build(self, filepaths):
        # Period (dùng chung cho tất cả)
        top = tk.Frame(self, padx=16, pady=10); top.pack(fill="x")
        tk.Label(top, text="Tháng / Kỳ:", font=("Arial", 10, "bold")).pack(side="left")
        self._period_var = tk.StringVar(value=datetime.now().strftime("%m/%Y"))
        tk.Entry(top, textvariable=self._period_var, width=12,
                 font=("Arial", 11)).pack(side="left", padx=8)
        tk.Label(top, text="(áp dụng cho tất cả file bên dưới)",
                 font=("Arial", 8), fg=C_GRAY).pack(side="left")

        # Table header
        hdr = tk.Frame(self, bg="#E5E7EB", pady=4); hdr.pack(fill="x", padx=12)
        tk.Label(hdr, text="Tên file", font=("Arial", 9, "bold"),
                 bg="#E5E7EB", width=46, anchor="w").pack(side="left", padx=6)
        tk.Label(hdr, text="Tên trạm (có thể sửa)", font=("Arial", 9, "bold"),
                 bg="#E5E7EB", width=20, anchor="w").pack(side="left")

        # Scrollable file rows
        frame = tk.Frame(self); frame.pack(fill="both", expand=True, padx=12, pady=2)
        cnv = tk.Canvas(frame, highlightthickness=0); cnv.pack(side="left", fill="both", expand=True)
        sb  = ttk.Scrollbar(frame, orient="vertical", command=cnv.yview); sb.pack(side="right", fill="y")
        sf  = tk.Frame(cnv)
        sf.bind("<Configure>", lambda e: cnv.configure(scrollregion=cnv.bbox("all")))
        cnv.create_window((0, 0), window=sf, anchor="nw")
        cnv.configure(yscrollcommand=sb.set)
        cnv.bind_all("<MouseWheel>", lambda e: cnv.yview_scroll(-1*(e.delta//120), "units"))

        for i, fp in enumerate(filepaths):
            bg = "#F9FAFB" if i % 2 == 0 else "white"
            row = tk.Frame(sf, bg=bg, pady=3); row.pack(fill="x", padx=2, pady=1)
            sv = tk.StringVar(value=self._auto_station(Path(fp).stem))
            self._station_vars.append(sv)
            tk.Label(row, text=Path(fp).name, width=48, anchor="w", bg=bg,
                     font=("Consolas", 8)).pack(side="left", padx=6)
            tk.Entry(row, textvariable=sv, width=16,
                     font=("Arial", 9)).pack(side="left", padx=4)

        # Buttons
        bf = tk.Frame(self, pady=10); bf.pack(fill="x", padx=12)
        tk.Button(bf, text=f"✓  Thêm {len(filepaths)} file vào danh sách",
                  font=("Arial", 10, "bold"), bg=C_GREEN, fg="white",
                  padx=16, pady=7, relief="flat",
                  command=self._confirm).pack(side="right", padx=4)
        tk.Button(bf, text="Hủy", font=("Arial", 10), bg=C_GRAY, fg="white",
                  padx=16, pady=7, relief="flat",
                  command=self.destroy).pack(side="right")

    def _confirm(self):
        period = self._period_var.get().strip()
        for fp, sv in zip(self._filepaths, self._station_vars):
            st = sv.get().strip()
            if not st: continue
            self.result.append({
                "filepath": fp, "station": st, "period": period,
                "display": f"[{st}]  {period}  —  {Path(fp).name}",
            })
        self.destroy()


# ─── Add File Dialog ──────────────────────────────────────────────────────────

class AddFileDialog(tk.Toplevel):
    """Popup: chọn file + nhập tên trạm + chọn tháng."""

    def __init__(self, parent):
        super().__init__(parent)
        self.title("Thêm file Excel")
        self.geometry("480x260")
        self.resizable(False, False)
        self.grab_set()
        self.result = None  # {"filepath", "station", "period"}

        tk.Label(self, text="Thêm file dữ liệu trạm", font=("Arial", 11, "bold"),
                 pady=10).pack()

        frm = tk.Frame(self, padx=20, pady=4)
        frm.pack(fill="x")

        # File path
        tk.Label(frm, text="File Excel:", font=("Arial", 9), anchor="w").grid(
            row=0, column=0, sticky="w", pady=6)
        self._path_var = tk.StringVar()
        tk.Entry(frm, textvariable=self._path_var, width=32, state="readonly",
                 font=("Consolas", 8)).grid(row=0, column=1, padx=4)
        tk.Button(frm, text="Chọn...", command=self._pick_file,
                  bg=C_GREEN, fg="white", font=("Arial", 9), padx=8).grid(row=0, column=2)

        # Station name
        tk.Label(frm, text="Tên trạm / Khu:", font=("Arial", 9), anchor="w").grid(
            row=1, column=0, sticky="w", pady=6)
        self._station_var = tk.StringVar()
        station_entry = tk.Entry(frm, textvariable=self._station_var, width=18,
                                 font=("Arial", 10))
        station_entry.grid(row=1, column=1, sticky="w", padx=4)
        tk.Label(frm, text="(ví dụ: A6, C6, E4, F, D4...)",
                 font=("Arial", 8), fg=C_GRAY).grid(row=1, column=2, sticky="w")

        # Period
        tk.Label(frm, text="Tháng / Kỳ:", font=("Arial", 9), anchor="w").grid(
            row=2, column=0, sticky="w", pady=6)
        self._period_var = tk.StringVar(value=datetime.now().strftime("%m/%Y"))
        tk.Entry(frm, textvariable=self._period_var, width=18,
                 font=("Arial", 10)).grid(row=2, column=1, sticky="w", padx=4)
        tk.Label(frm, text="(ví dụ: 03/2026, 5.2026...)",
                 font=("Arial", 8), fg=C_GRAY).grid(row=2, column=2, sticky="w")

        # Buttons
        btn = tk.Frame(self, pady=12)
        btn.pack()
        tk.Button(btn, text="✓  Thêm vào danh sách", command=self._confirm,
                  font=("Arial", 10, "bold"), bg=C_GREEN, fg="white",
                  padx=16, pady=7, relief="flat").pack(side="left", padx=6)
        tk.Button(btn, text="Hủy", command=self.destroy,
                  font=("Arial", 10), bg=C_GRAY, fg="white",
                  padx=16, pady=7, relief="flat").pack(side="left")

    def _pick_file(self):
        p = filedialog.askopenfilename(
            title="Chọn file Excel của trạm",
            filetypes=[("Excel", "*.xlsx *.xls")])
        if p:
            self._path_var.set(p)
            # Tự gợi ý tên trạm từ tên file
            stem = Path(p).stem
            # Tìm pattern như A6, C6, E4, KHU_E, etc.
            m = re.search(r"(?:khu[_\s]*)?([A-Za-z]\d+|[A-Za-z]+\s*\d*)", stem, re.IGNORECASE)
            if m and not self._station_var.get():
                self._station_var.set(m.group(1).strip().upper())

    def _confirm(self):
        fp = self._path_var.get().strip()
        st = self._station_var.get().strip()
        pr = self._period_var.get().strip()
        if not fp:
            messagebox.showwarning("Thiếu file", "Vui lòng chọn file Excel.", parent=self)
            return
        if not st:
            messagebox.showwarning("Thiếu tên trạm", "Vui lòng nhập tên trạm/khu.", parent=self)
            return
        self.result = {"filepath": fp, "station": st, "period": pr}
        self.destroy()

# ─── Column Mapping Dialog ────────────────────────────────────────────────────

class ColMappingDialog(tk.Toplevel):
    """
    Hiển thị cấu trúc phát hiện được từ 1 file Excel và cho phép user chỉnh.
    Tabs: Cột bệnh nhân | Dòng thuốc | Xem trước
    """
    def __init__(self, parent, filepath: str, rows: list, struct: dict):
        super().__init__(parent)
        self.title(f"Kiểm tra cấu trúc — {Path(filepath).name}")
        self.geometry("900x540")
        self.resizable(True, True)
        self.grab_set()
        self.confirmed   = False
        self.apply_to_all = tk.BooleanVar(value=True)
        self.result_struct: dict = {}
        self._filepath = filepath
        self._rows     = rows
        self._struct   = struct
        self._col_vars: dict[int, tk.StringVar] = {}   # col_idx → field label var
        self._row_vars: dict[int, tk.StringVar] = {}   # row_idx → med field label var
        self._build()
        self.protocol("WM_DELETE_WINDOW", self._cancel)

    # ── UI ────────────────────────────────────────────────────────────────────
    def _build(self):
        s = self._struct
        col_field_inv = {v: k for k, v in s["col_map"].items()}   # col_idx → field
        row_field_inv = {v: k for k, v in s["med_header_row_idx"].items()}  # row_idx → field

        bar = tk.Frame(self, bg=C_GREEN); bar.pack(fill="x")
        tk.Label(bar, text=f"  Kiểm tra & chỉnh cấu trúc file — {Path(self._filepath).name}",
                 font=("Arial", 10, "bold"), fg="white", bg=C_GREEN, pady=6).pack(side="left")

        nb = ttk.Notebook(self); nb.pack(fill="both", expand=True, padx=8, pady=4)

        f1 = tk.Frame(nb); nb.add(f1, text="  Cột thông tin bệnh nhân  ")
        self._tab_patient_cols(f1, col_field_inv)

        f2 = tk.Frame(nb); nb.add(f2, text="  Dòng thông tin thuốc  ")
        self._tab_med_rows(f2, row_field_inv)

        f3 = tk.Frame(nb); nb.add(f3, text="  Xem trước dữ liệu  ")
        self._tab_preview(f3)

        bf = tk.Frame(self, pady=6); bf.pack(fill="x", padx=12)
        ttk.Checkbutton(bf, text="Áp dụng cấu trúc này cho tất cả file còn lại",
                        variable=self.apply_to_all).pack(side="left", padx=6)
        tk.Button(bf, text="✓  Xác nhận", font=("Arial", 10, "bold"),
                  bg=C_GREEN, fg="white", padx=16, pady=6, relief="flat",
                  command=self._confirm).pack(side="right", padx=4)
        tk.Button(bf, text="✕  Hủy nhập", font=("Arial", 10),
                  bg=C_RED, fg="white", padx=16, pady=6, relief="flat",
                  command=self._cancel).pack(side="right")

    def _scrollable_frame(self, parent):
        cnv = tk.Canvas(parent, highlightthickness=0); cnv.pack(fill="both", expand=True, padx=8)
        sb  = ttk.Scrollbar(cnv, orient="vertical", command=cnv.yview)
        sf  = tk.Frame(cnv, bg="white")
        sf.bind("<Configure>", lambda e: cnv.configure(scrollregion=cnv.bbox("all")))
        cnv.create_window((0, 0), window=sf, anchor="nw")
        cnv.configure(yscrollcommand=sb.set); sb.pack(side="right", fill="y")
        cnv.bind_all("<MouseWheel>", lambda e: cnv.yview_scroll(-1*(e.delta//120), "units"))
        return sf

    def _tab_patient_cols(self, parent, col_field_inv):
        tk.Label(parent,
                 text="  Mỗi cột trong file → gán trường bệnh nhân. Cột nhãn thuốc (I) và cột thuốc (J+) để Bỏ qua.",
                 font=("Arial", 8), fg="#6B7280", bg="#F9FAFB", pady=3).pack(fill="x")

        th = tk.Frame(parent, bg="#E5E7EB", pady=3); th.pack(fill="x", padx=6, pady=(0, 2))
        for txt, w in [("Cột", 6), ("Nội dung header", 32), ("Trường gán", 24)]:
            tk.Label(th, text=txt, font=("Arial", 8, "bold"), bg="#E5E7EB",
                     width=w, anchor="w").pack(side="left", padx=4)

        sf = self._scrollable_frame(parent)
        opts    = list(PATIENT_FIELD_LABELS.values())
        f2label = {f: l for f, l in PATIENT_FIELD_LABELS.items()}
        s = self._struct
        hdr_row = self._rows[s["header_row_idx"]] if s["header_row_idx"] < len(self._rows) else []
        lc = s["label_col"]; msc = s["med_start_col"]

        n_show = min(msc + 3, len(hdr_row), 30)
        for ci in range(n_show):
            cell_text = str(hdr_row[ci] if ci < len(hdr_row) else "").strip()
            col_letter = get_column_letter(ci + 1)
            detected   = col_field_inv.get(ci, "_ignore")

            if ci == lc:
                bg = "#FEF9C3"; tag = "← nhãn thuốc"
            elif ci >= msc:
                bg = "#F0FDF4"; tag = "← cột thuốc"
            else:
                bg = "white" if ci % 2 == 0 else "#F9FAFB"; tag = ""

            row_fr = tk.Frame(sf, bg=bg, pady=2); row_fr.pack(fill="x", padx=2, pady=1)
            tk.Label(row_fr, text=col_letter, width=6, font=("Consolas", 9, "bold"),
                     bg=bg, anchor="center").pack(side="left", padx=4)
            disp = (cell_text[:30] + "…" if len(cell_text) > 30 else cell_text) + ("  " + tag if tag else "")
            tk.Label(row_fr, text=disp, width=32, font=("Consolas", 9),
                     bg=bg, anchor="w").pack(side="left", padx=2)

            if ci >= lc:
                tk.Label(row_fr, text="(không gán)", font=("Arial", 8), fg="#9CA3AF",
                         bg=bg, width=24, anchor="w").pack(side="left")
            else:
                var = tk.StringVar(value=f2label.get(detected, "(Bỏ qua)"))
                self._col_vars[ci] = var
                ttk.Combobox(row_fr, textvariable=var, values=opts,
                             width=24, state="readonly").pack(side="left", padx=4)

    def _tab_med_rows(self, parent, row_field_inv):
        tk.Label(parent,
                 text="  Mỗi dòng header → gán trường thuốc. Nhãn đọc từ cột I (cột nhãn).",
                 font=("Arial", 8), fg="#6B7280", bg="#F9FAFB", pady=3).pack(fill="x")

        th = tk.Frame(parent, bg="#E5E7EB", pady=3); th.pack(fill="x", padx=6, pady=(0, 2))
        for txt, w in [("Dòng", 6), ("Nhãn (cột I)", 32), ("Trường gán", 24)]:
            tk.Label(th, text=txt, font=("Arial", 8, "bold"), bg="#E5E7EB",
                     width=w, anchor="w").pack(side="left", padx=4)

        sf = self._scrollable_frame(parent)
        opts    = list(MED_FIELD_LABELS.values())
        f2label = {f: l for f, l in MED_FIELD_LABELS.items()}
        s = self._struct; lc = s["label_col"]

        for ri in range(s["actual_data_start"]):
            row = self._rows[ri]
            cells = [str(c or "").strip() for c in row]
            lbl   = cells[lc] if lc < len(cells) else ""
            if not lbl:
                for ci in range(min(lc, len(cells))):
                    if cells[ci]: lbl = cells[ci]; break

            detected = row_field_inv.get(ri, "_ignore")
            bg = "#F0FDF4" if ri % 2 == 0 else "white"
            row_fr = tk.Frame(sf, bg=bg, pady=2); row_fr.pack(fill="x", padx=2, pady=1)
            tk.Label(row_fr, text=f"Dòng {ri+1}", width=6, font=("Consolas", 9, "bold"),
                     bg=bg, anchor="center").pack(side="left", padx=4)
            disp = lbl[:40] + "…" if len(lbl) > 40 else lbl
            tk.Label(row_fr, text=disp, width=32, font=("Consolas", 9),
                     bg=bg, anchor="w").pack(side="left", padx=2)
            var = tk.StringVar(value=f2label.get(detected, "(Bỏ qua)"))
            self._row_vars[ri] = var
            ttk.Combobox(row_fr, textvariable=var, values=opts,
                         width=24, state="readonly").pack(side="left", padx=4)

    def _tab_preview(self, parent):
        tk.Label(parent, text="  5 dòng đầu dữ liệu bệnh nhân:",
                 font=("Arial", 9), fg="#6B7280", pady=3).pack(anchor="w", padx=8)
        txt = tk.Text(parent, height=14, font=("Consolas", 8), wrap="none",
                      bg="#FAFAFA", relief="flat")
        sbx = ttk.Scrollbar(parent, orient="horizontal", command=txt.xview)
        sby = ttk.Scrollbar(parent, orient="vertical",   command=txt.yview)
        txt.configure(xscrollcommand=sbx.set, yscrollcommand=sby.set)
        sbx.pack(side="bottom", fill="x"); sby.pack(side="right", fill="y")
        txt.pack(fill="both", expand=True, padx=8, pady=4)
        ads = self._struct["actual_data_start"]
        for row in self._rows[ads:ads + 5]:
            line = "  ".join(f"{str(c or '')[:14]:14}" for c in row[:18])
            txt.insert("end", line + "\n")
        txt.config(state="disabled")

    # ── Confirm / Cancel ──────────────────────────────────────────────────────
    def _confirm(self):
        l2f_pat = {l: f for f, l in PATIENT_FIELD_LABELS.items()}
        l2f_med = {l: f for f, l in MED_FIELD_LABELS.items()}

        col_map = {}
        for ci, var in self._col_vars.items():
            f = l2f_pat.get(var.get(), "_ignore")
            if f != "_ignore": col_map[f] = ci

        med_row_idx = {}
        for ri, var in self._row_vars.items():
            f = l2f_med.get(var.get(), "_ignore")
            if f != "_ignore": med_row_idx[f] = ri

        last_pat = max(col_map.values()) if col_map else self._struct["label_col"] - 1
        self.result_struct = dict(self._struct)
        self.result_struct["col_map"]            = col_map
        self.result_struct["med_header_row_idx"] = med_row_idx
        self.result_struct["label_col"]          = last_pat + 1
        self.result_struct["med_start_col"]      = last_pat + 2
        self.confirmed = True
        self.destroy()

    def _cancel(self):
        self.confirmed = False; self.destroy()


# ─── Disease Group Normalization Dialog ────────────────────────────────────────

class DiseaseGroupDialog(tk.Toplevel):
    """
    Liệt kê tất cả nhóm bệnh tìm thấy trong DB.
    User có thể map một nhóm → nhóm khác để gộp lại khi tính báo cáo.
    Mặc định: giữ nguyên tất cả.
    """
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Chuẩn hóa nhóm bệnh — 疾病分类整理")
        self.geometry("760x480")
        self.resizable(True, True)
        self.grab_set()
        self.confirmed = False
        self.result: dict[str, str] = {}   # raw_group → canonical_group (same if no change)
        self._vars: dict[str, tk.StringVar] = {}
        self._groups: list[str] = []
        self._build()
        self.protocol("WM_DELETE_WINDOW", self._cancel)

    def _build(self):
        conn = get_db()
        try:
            rows = conn.execute(
                "SELECT DISTINCT disease_group, COUNT(*) AS cnt "
                "FROM encounters WHERE disease_group != '' "
                "GROUP BY disease_group ORDER BY cnt DESC"
            ).fetchall()
        finally:
            conn.close()

        self._groups = [r["disease_group"] for r in rows]
        counts       = {r["disease_group"]: r["cnt"] for r in rows}

        bar = tk.Frame(self, bg=C_GREEN); bar.pack(fill="x")
        tk.Label(bar, text=f"  Chuẩn hóa nhóm bệnh ({len(self._groups)} nhóm tìm thấy)",
                 font=("Arial", 10, "bold"), fg="white", bg=C_GREEN, pady=6).pack(side="left")

        info = tk.Frame(self, bg="#FEF9C3", pady=3); info.pack(fill="x")
        tk.Label(info, text="  Chọn nhóm đích để gộp, hoặc giữ nguyên. Thay đổi chỉ áp dụng cho báo cáo (không xóa dữ liệu gốc).",
                 font=("Arial", 8), bg="#FEF9C3", fg="#92400E").pack(side="left")

        th = tk.Frame(self, bg="#E5E7EB", pady=3); th.pack(fill="x", padx=8, pady=(4, 0))
        for txt, w in [("Nhóm gốc trong DB", 30), ("Số ca", 8), ("→ Map thành", 28)]:
            tk.Label(th, text=txt, font=("Arial", 8, "bold"), bg="#E5E7EB",
                     width=w, anchor="w").pack(side="left", padx=4)

        cnv = tk.Canvas(self, highlightthickness=0); cnv.pack(fill="both", expand=True, padx=8)
        sb  = ttk.Scrollbar(cnv, orient="vertical", command=cnv.yview)
        sf  = tk.Frame(cnv)
        sf.bind("<Configure>", lambda e: cnv.configure(scrollregion=cnv.bbox("all")))
        cnv.create_window((0, 0), window=sf, anchor="nw")
        cnv.configure(yscrollcommand=sb.set); sb.pack(side="right", fill="y")
        cnv.bind_all("<MouseWheel>", lambda e: cnv.yview_scroll(-1*(e.delta//120), "units"))

        all_opts = ["(giữ nguyên)"] + self._groups
        for i, g in enumerate(self._groups):
            bg  = "#F9FAFB" if i % 2 == 0 else "white"
            row = tk.Frame(sf, bg=bg, pady=3); row.pack(fill="x", padx=2, pady=1)
            tk.Label(row, text=g, width=30, anchor="w", bg=bg,
                     font=("Consolas", 9)).pack(side="left", padx=4)
            tk.Label(row, text=str(counts[g]), width=8, anchor="center", bg=bg,
                     font=("Arial", 9)).pack(side="left", padx=2)
            var = tk.StringVar(value="(giữ nguyên)")
            self._vars[g] = var
            ttk.Combobox(row, textvariable=var, values=all_opts,
                         width=28, state="readonly").pack(side="left", padx=4)

        bf = tk.Frame(self, pady=8); bf.pack(fill="x", padx=12)
        tk.Button(bf, text="✓  Áp dụng", font=("Arial", 10, "bold"),
                  bg=C_GREEN, fg="white", padx=16, pady=6, relief="flat",
                  command=self._confirm).pack(side="right", padx=4)
        tk.Button(bf, text="✕  Hủy", font=("Arial", 10),
                  bg=C_RED, fg="white", padx=16, pady=6, relief="flat",
                  command=self._cancel).pack(side="right")

    def _confirm(self):
        for g, var in self._vars.items():
            sel = var.get()
            self.result[g] = g if sel == "(giữ nguyên)" else sel
        self.confirmed = True; self.destroy()

    def _cancel(self):
        self.confirmed = False; self.destroy()


# ─── Medicine Mapping Dialog ───────────────────────────────────────────────────

class MappingDialog(tk.Toplevel):
    def __init__(self, parent, unmatched: list[dict]):
        super().__init__(parent)
        self.title("Xác nhận tên thuốc — 药品名称匹配")
        self.geometry("900x520")
        self.resizable(True, True)
        self.grab_set()
        self.confirmed = False
        self.result: dict[str, tuple] = {}
        self._vars: dict[str, tuple] = {}
        self._rows_meta: list[dict] = []
        self._threshold = tk.IntVar(value=70)
        self._build(unmatched)
        self._threshold.trace_add("write", lambda *_: self._apply_threshold())
        self.protocol("WM_DELETE_WINDOW", self._cancel)

    # ── Compute default for one item at given threshold (pct = 0-100) ──────────
    def _default_for(self, item: dict, pct: int) -> tuple:
        """Returns (action, name, reason_txt, reason_fg)"""
        zero = item.get("all_zero", False)
        sug  = item["suggestions"]   # [(id, name, score), ...]
        cr   = item.get("cross_match")  # (other_name, score) | None
        en   = item["excel_name"]

        if zero:
            return "skip", en, "Tất cả = 0", C_GRAY

        thresh = pct / 100
        sug_ok = [s for s in sug if s[2] >= thresh]
        if sug_ok:
            best = sug_ok[0]
            fg   = "#16A34A" if best[2] >= 0.8 else C_AMBER
            return "map", best[1], f"DB {best[2]:.0%}", fg

        cr_ok = cr if (cr and cr[1] >= thresh) else None
        if cr_ok:
            return "map", cr_ok[0], f"Giống {cr_ok[1]:.0%}", C_AMBER

        return "new", en, "Mới hoàn toàn", C_BLUE

    # ── Re-apply threshold to all non-manually-set rows ───────────────────────
    def _apply_threshold(self):
        pct = self._threshold.get()
        for m in self._rows_meta:
            if m["manually_set"]:
                continue
            action, name, rtxt, rfg = self._default_for(m["item"], pct)
            m["av"].set(action)
            m["nv"].set(name)
            m["reason_lbl"].config(text=rtxt, fg=rfg)

    def _build(self, unmatched):
        # ── Header bar ───────────────────────────────────────────────────────
        top = tk.Frame(self, bg=C_GREEN); top.pack(fill="x")
        tk.Label(top, text=f"  {len(unmatched)} tên thuốc chưa khớp — vui lòng xác nhận:",
                 font=("Arial", 10, "bold"), fg="white", bg=C_GREEN, pady=8).pack(side="left")

        # ── Legend ───────────────────────────────────────────────────────────
        legend = tk.Frame(self, bg="#FEF9C3", pady=3); legend.pack(fill="x")
        tk.Label(legend,
                 text=("  map = ghép tên đã có  │  new = thêm mới  "
                       "│  skip = bỏ qua (= 0 hoặc trùng)"),
                 font=("Arial", 8), bg="#FEF9C3", fg="#92400E").pack(side="left")

        # ── Threshold slider ──────────────────────────────────────────────────
        tbar = tk.Frame(self, bg="#EFF6FF", pady=5); tbar.pack(fill="x", padx=8, pady=(4, 0))
        tk.Label(tbar, text="Ngưỡng gợi ý map:", font=("Arial", 9, "bold"),
                 bg="#EFF6FF").pack(side="left", padx=6)
        self._thr_lbl = tk.Label(tbar, text="70%", font=("Arial", 9, "bold"),
                                  bg="#EFF6FF", fg=C_BLUE, width=5)
        self._thr_lbl.pack(side="left")

        def _on_scale(v):
            self._thr_lbl.config(text=f"{int(float(v))}%")

        tk.Scale(tbar, from_=50, to=95, orient="horizontal",
                 variable=self._threshold, showvalue=False, length=220,
                 bg="#EFF6FF", troughcolor="#BFDBFE", activebackground=C_BLUE,
                 command=_on_scale).pack(side="left", padx=4)
        tk.Label(tbar, text="← dễ map hơn  |  khắt khe hơn →  "
                             "(row đã sửa tay sẽ không đổi)",
                 font=("Arial", 8), fg="#6B7280", bg="#EFF6FF").pack(side="left", padx=8)
        tk.Button(tbar, text="Reset tay", font=("Arial", 8),
                  relief="flat", bg="#DBEAFE", fg="#1D4ED8", padx=6,
                  command=self._reset_manual).pack(side="right", padx=6)

        # ── Column headers ────────────────────────────────────────────────────
        th = tk.Frame(self, bg="#E5E7EB", pady=3); th.pack(fill="x", padx=8, pady=(4, 0))
        for txt, w in [("Tên trong file", 30), ("Đề xuất", 8),
                       ("Tên chuẩn (chọn hoặc tự nhập)", 36), ("Lý do", 12)]:
            tk.Label(th, text=txt, font=("Arial", 8, "bold"),
                     bg="#E5E7EB", width=w, anchor="w").pack(side="left", padx=3)

        # ── Scrollable list ───────────────────────────────────────────────────
        cnv = tk.Canvas(self, highlightthickness=0); cnv.pack(fill="both", expand=True, padx=8)
        sb  = ttk.Scrollbar(cnv, orient="vertical", command=cnv.yview)
        sf  = tk.Frame(cnv)
        sf.bind("<Configure>", lambda e: cnv.configure(scrollregion=cnv.bbox("all")))
        cnv.create_window((0, 0), window=sf, anchor="nw")
        cnv.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        cnv.bind_all("<MouseWheel>", lambda e: cnv.yview_scroll(-1*(e.delta//120), "units"))

        pct = self._threshold.get()
        for i, item in enumerate(unmatched):
            en  = item["excel_name"]
            sug = item["suggestions"]
            cr  = item.get("cross_match")

            action, name, rtxt, rfg = self._default_for(item, pct)

            # Row background by category (stays fixed; threshold only changes label/vars)
            zero = item.get("all_zero", False)
            if zero:
                row_bg = "#F3F4F6"
            elif action == "map" and sug and sug[0][2] >= pct/100:
                row_bg = "#F0FDF4" if i % 2 == 0 else "#E6FAF0"
            elif action == "map":
                row_bg = "#FFFBEB" if i % 2 == 0 else "#FEF9C3"
            else:
                row_bg = "#F9FAFB" if i % 2 == 0 else "white"

            av = tk.StringVar(value=action)
            nv = tk.StringVar(value=name)
            self._vars[en] = (av, nv)

            row = tk.Frame(sf, bg=row_bg, pady=3); row.pack(fill="x", padx=2, pady=1)

            tk.Label(row, text=en, width=32, anchor="w", bg=row_bg,
                     font=("Consolas", 9)).pack(side="left", padx=3)

            meta = {"item": item, "av": av, "nv": nv, "manually_set": False,
                    "reason_lbl": None}
            self._rows_meta.append(meta)

            act_cb = ttk.Combobox(row, textvariable=av, values=["map", "new", "skip"],
                                  width=7, state="readonly")
            act_cb.pack(side="left", padx=3)
            act_cb.bind("<<ComboboxSelected>>",
                        lambda e, m=meta: m.__setitem__("manually_set", True))

            opts = [s[1] for s in sug]
            if cr and cr[0] not in opts: opts.insert(0, cr[0])
            if en not in opts: opts.append(en)
            name_cb = ttk.Combobox(row, textvariable=nv, values=opts, width=36)
            name_cb.pack(side="left", padx=3)
            name_cb.bind("<<ComboboxSelected>>",
                         lambda e, m=meta: m.__setitem__("manually_set", True))

            rlbl = tk.Label(row, text=rtxt, fg=rfg, bg=row_bg,
                            font=("Arial", 8, "bold"), width=14, anchor="w")
            rlbl.pack(side="left")
            meta["reason_lbl"] = rlbl

        # ── Bottom buttons ────────────────────────────────────────────────────
        bf = tk.Frame(self, pady=8); bf.pack(fill="x", padx=12)
        tk.Button(bf, text="✓  Xác nhận", font=("Arial", 10, "bold"),
                  bg=C_GREEN, fg="white", padx=18, pady=7, relief="flat",
                  command=self._confirm).pack(side="right", padx=4)
        tk.Button(bf, text="✕  Hủy", font=("Arial", 10),
                  bg=C_RED, fg="white", padx=18, pady=7, relief="flat",
                  command=self._cancel).pack(side="right")

    def _reset_manual(self):
        """Clear all manual-set flags and re-apply threshold."""
        for m in self._rows_meta:
            m["manually_set"] = False
        self._apply_threshold()

    def _confirm(self):
        for en, (av, nv) in self._vars.items():
            self.result[en] = (av.get(), nv.get().strip())
        self.confirmed = True; self.destroy()

    def _cancel(self):
        self.confirmed = False; self.destroy()

# ─── Main App ─────────────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("GVC Medical Import Tool")
        self.geometry("1000x680")
        self.minsize(800, 560)
        self.configure(bg=C_BG)
        # {"filepath", "station", "period", "display": str}
        self._files: list[dict] = []
        self._build_ui()
        self._refresh_stats()

    def _build_ui(self):
        # Title bar
        bar = tk.Frame(self, bg=C_GREEN); bar.pack(fill="x")
        tk.Label(bar, text="  GVC Medical Import Tool",
                 font=("Arial", 13, "bold"), fg="white", bg=C_GREEN, pady=9).pack(side="left")
        tk.Label(bar, text="Nhập & tổng hợp dữ liệu y tế",
                 font=("Arial", 9), fg="#BBF7D0", bg=C_GREEN).pack(side="left", padx=4)
        self._stats_var = tk.StringVar(value="")
        tk.Label(bar, textvariable=self._stats_var, font=("Arial", 8),
                 fg="#A7F3D0", bg=C_GREEN).pack(side="right", padx=12)

        # Actions
        act = tk.Frame(self, bg=C_BG, pady=8); act.pack(fill="x", padx=14)
        self._btn(act, "📋  Tạo Template",      self._template,   C_BLUE).pack(side="left", padx=3)
        self._btn(act, "➕  Thêm nhiều file",  self._add_files,  C_GREEN).pack(side="left", padx=3)
        self._btn(act, "✕  Xóa danh sách",    self._clear,      C_GRAY).pack(side="left", padx=3)

        ttk.Separator(self, orient="horizontal").pack(fill="x", padx=14, pady=2)

        # Main area
        main = tk.Frame(self, bg=C_BG); main.pack(fill="both", expand=True, padx=14, pady=4)

        left = tk.LabelFrame(main, text=" Danh sách file sẽ nhập ",
                              font=("Arial",9), bg=C_BG, padx=6, pady=6)
        left.pack(side="left", fill="both", expand=True, padx=(0,6))
        self._flist = tk.Listbox(left, font=("Consolas", 8), selectmode="extended",
                                  bg="white", relief="flat",
                                  highlightthickness=1, highlightcolor=COLOR_LIGHT if False else "#D1FAE5")
        sbf = ttk.Scrollbar(left, orient="vertical", command=self._flist.yview)
        self._flist.configure(yscrollcommand=sbf.set)
        sbf.pack(side="right", fill="y"); self._flist.pack(fill="both", expand=True)

        right = tk.LabelFrame(main, text=" Nhật ký ",
                               font=("Arial",9), bg=C_BG, padx=6, pady=6)
        right.pack(side="right", fill="both", expand=True)
        self._log = tk.Text(right, font=("Consolas", 8), state="disabled",
                            bg="#0F172A", fg="#A3E635", relief="flat", wrap="word")
        sbl = ttk.Scrollbar(right, orient="vertical", command=self._log.yview)
        self._log.configure(yscrollcommand=sbl.set)
        sbl.pack(side="right", fill="y"); self._log.pack(fill="both", expand=True)

        # Bottom
        bot = tk.Frame(self, bg=C_BG, pady=10); bot.pack(fill="x", padx=14)
        self._btn(bot, "🚀  Nhập dữ liệu vào DB",            self._import,            C_GREEN,   large=True).pack(side="left", padx=3)
        self._btn(bot, "📋  Xuất danh sách BN (File 1)",      self._export_patient,    "#059669", large=True).pack(side="left", padx=3)
        self._btn(bot, "📊  Xuất báo cáo mặt bệnh (File 2)",  self._export_disease,    "#7C3AED", large=True).pack(side="left", padx=3)
        self._btn(bot, "💊  Xuất báo cáo thuốc (File 3)",     self._export_medicine,   "#0369A1", large=True).pack(side="left", padx=3)
        self._btn(bot, "🏷  Nhóm bệnh",                       self._disease_group_dlg, "#B45309").pack(side="left", padx=3)
        self._btn(bot, "🗑  Reset DB", self._reset, C_RED).pack(side="right", padx=3)

        self._status = tk.StringVar(value="Sẵn sàng")
        tk.Label(self, textvariable=self._status, font=("Arial", 8), fg=C_GRAY,
                 bg="#F3F4F6", relief="sunken", anchor="w", padx=8, pady=3).pack(fill="x", side="bottom")

    def _btn(self, p, txt, cmd, color, large=False):
        return tk.Button(p, text=txt, command=cmd,
                         font=("Arial", 10 if large else 9, "bold"),
                         bg=color, fg="white",
                         padx=14 if large else 10, pady=7 if large else 5,
                         relief="flat", cursor="hand2",
                         activebackground=color, activeforeground="white")

    def _log_msg(self, msg):
        self._log.configure(state="normal")
        self._log.insert("end", f"[{datetime.now().strftime('%H:%M:%S')}] {msg}\n")
        self._log.see("end"); self._log.configure(state="disabled")
        self._status.set(msg[:90]); self.update_idletasks()

    def _refresh_stats(self):
        conn = get_db(); s = db_stats(conn); conn.close()
        self._stats_var.set(
            f"DB: {s['encounters']} ca khám  |  {s['medicines']} thuốc  |  {s['stations']} trạm"
        )

    # ── Actions ─────────────────────────────────────────────────────────────

    def _template(self):
        p = filedialog.asksaveasfilename(title="Lưu template",
                                          defaultextension=".xlsx",
                                          filetypes=[("Excel","*.xlsx")],
                                          initialfile="Template_Tram_Y_Te.xlsx")
        if not p: return
        try:
            generate_template(p)
            self._log_msg(f"✅ Template: {Path(p).name}")
            messagebox.showinfo("Thành công", f"Template đã tạo:\n{p}")
        except Exception as e:
            self._log_msg(f"❌ {e}"); messagebox.showerror("Lỗi", str(e))

    def _add_files(self):
        paths = filedialog.askopenfilenames(
            title="Chọn các file Excel của các trạm",
            filetypes=[("Excel", "*.xlsx *.xls")])
        if not paths: return
        dlg = MultiFileDialog(self, list(paths))
        self.wait_window(dlg)
        if not dlg.result: return

        # Kiểm tra file nào đã được nhập vào DB rồi
        conn = get_db()
        try:
            already_done, to_add = [], []
            for r in dlg.result:
                fn = Path(r["filepath"]).name
                row = conn.execute(
                    "SELECT imported_at FROM imported_files "
                    "WHERE filename=? AND station=? AND period=?",
                    (fn, r["station"], r["period"])
                ).fetchone()
                if row:
                    already_done.append((r, row["imported_at"][:16].replace("T", " ")))
                else:
                    to_add.append(r)
        finally:
            conn.close()

        if already_done:
            lines = "\n".join(
                f"  • [{r['station']}] {Path(r['filepath']).name}  (đã nhập: {ts})"
                for r, ts in already_done
            )
            ans = messagebox.askyesnocancel(
                "File đã được nhập",
                f"Các file sau ĐÃ có trong database:\n{lines}\n\n"
                "• Bấm Yes → Bỏ qua, chỉ thêm file mới\n"
                "• Bấm No  → Thêm tất cả (sẽ ghi đè dữ liệu cũ khi nhập)\n"
                "• Cancel  → Hủy",
                parent=self
            )
            if ans is None:      # Cancel
                return
            elif ans:            # Yes → chỉ thêm file mới
                pass             # to_add đã đúng
            else:                # No → thêm tất cả (force overwrite)
                to_add = dlg.result

        for r in to_add:
            self._files.append(r)
            self._flist.insert("end", r["display"])
            self._log_msg(f"✅ Thêm: [{r['station']}]  {r['period']}  —  {Path(r['filepath']).name}")

    def _clear(self):
        self._files.clear(); self._flist.delete(0, "end")
        self._log_msg("🗑 Đã xóa danh sách file.")

    def _import(self):
        if not self._files:
            messagebox.showwarning("Chưa có file", "Vui lòng thêm file trạm trước."); return

        conn = get_db()
        total_enc = 0
        try:
            # Pass 0: detect structure + ColMappingDialog
            self._log_msg(f"🔍 Phân tích cấu trúc {len(self._files)} file...")
            confirmed_struct = None   # confirmed by user via dialog
            apply_all        = False
            file_raw: list[tuple] = []  # (fi, rows, struct)

            for fi in self._files:
                rows, err = _load_excel_rows(fi["filepath"], fi["period"])
                if err:
                    self._log_msg(f"   ❌ [{fi['station']}] {err}"); continue
                struct = _detect_structure(rows)
                if struct.get("error"):
                    self._log_msg(f"   ❌ [{fi['station']}] {struct['error']}"); continue
                file_raw.append((fi, rows, struct))

            if not file_raw:
                self._log_msg("❌ Không có file hợp lệ."); return

            # Show ColMappingDialog for first file (or each if apply_all=False)
            for idx, (fi, rows, auto_struct) in enumerate(file_raw):
                if apply_all and confirmed_struct is not None:
                    # Reuse confirmed struct but re-run detect for actual_data_start
                    s = dict(confirmed_struct)
                    s["actual_data_start"] = auto_struct["actual_data_start"]
                    file_raw[idx] = (fi, rows, s)
                    continue
                dlg = ColMappingDialog(self, fi["filepath"], rows, auto_struct)
                self.wait_window(dlg)
                if not dlg.confirmed:
                    self._log_msg("❌ Người dùng hủy."); return
                confirmed_struct = dlg.result_struct
                apply_all        = dlg.apply_to_all.get()
                file_raw[idx]    = (fi, rows, confirmed_struct)

            # Pass 1: parse với confirmed structs + thu thập thuốc chưa biết
            self._log_msg(f"📖 Đọc {len(file_raw)} file...")
            file_results = []
            all_unmatched: dict[str, list] = {}
            all_med_stats: dict[str, float] = {}

            for fi, rows, struct in file_raw:
                self._log_msg(f"   [{fi['station']}] {Path(fi['filepath']).name}")
                res = _parse_from_struct(rows, struct, fi["station"], fi["period"])
                if res.get("error"):
                    self._log_msg(f"   ❌ {res['error']}"); continue
                self._log_msg(
                    f"   → {len(res['encounters'])} ca | {len(res['medicines'])} thuốc"
                    f"  [col J={res.get('med_start_col','?')}, "
                    f"data row {res.get('actual_data_start','?')+1}]"
                )
                file_results.append((fi, res))
                n_auto = 0
                for med in res["medicines"]:
                    n = med["name"]
                    if not n: continue
                    total = (med.get("beg_stock", 0) + med.get("received", 0)
                             + med.get("issued", 0) + med.get("used_header", 0))
                    all_med_stats[n] = all_med_stats.get(n, 0) + total
                    if find_med(conn, n):
                        n_auto += 1
                        continue
                    if n not in all_unmatched:
                        all_unmatched[n] = suggest_med(conn, n)
                if n_auto:
                    self._log_msg(f"  ↳ {n_auto} thuốc đã có trong DB, tự động map")

            if not file_results:
                self._log_msg("❌ Không có file hợp lệ."); return

            # Build unit lookup từ tất cả file đã parse (last-write-wins per medicine name)
            excel_units: dict[str, str] = {}
            for _fi, _res in file_results:
                for med in _res["medicines"]:
                    u = str(med.get("unit", "") or "").strip()
                    if u:
                        excel_units[med["name"]] = u

            # Pass 2: mapping dialog
            final_map: dict[str, int | None] = {}
            if all_unmatched:
                self._log_msg(f"⚠️ {len(all_unmatched)} tên thuốc cần xác nhận...")

                # Tính cross-match: tên thuốc chưa biết giống nhau
                unames = list(all_unmatched.keys())
                cross_best: dict[str, tuple] = {}
                for i, n1 in enumerate(unames):
                    for n2 in unames[i+1:]:
                        s = difflib.SequenceMatcher(None, norm(n1), norm(n2)).ratio()
                        if norm(n1) in norm(n2) or norm(n2) in norm(n1):
                            s = max(s, 0.82)
                        if s >= 0.62:
                            if n1 not in cross_best or s > cross_best[n1][1]:
                                cross_best[n1] = (n2, s)
                            if n2 not in cross_best or s > cross_best[n2][1]:
                                cross_best[n2] = (n1, s)

                # Build enriched list + smart defaults
                def _sort_key(item):
                    db  = item["suggestions"]
                    cr  = item["cross_match"]
                    # Nhóm theo tên canonical để thuốc giống nhau đứng cạnh nhau
                    canonical = (db[0][1] if db else cr[0] if cr else item["excel_name"])
                    return norm(canonical)

                unmatched_list = []
                for k, db_sug in all_unmatched.items():
                    cr   = cross_best.get(k)
                    zero = (all_med_stats.get(k, 0) == 0)
                    unmatched_list.append({
                        "excel_name":  k,
                        "suggestions": db_sug,
                        "cross_match": cr,
                        "all_zero":    zero,
                    })
                unmatched_list.sort(key=_sort_key)

                dlg = MappingDialog(self, unmatched_list)
                self.wait_window(dlg)
                if not dlg.confirmed:
                    self._log_msg("❌ Người dùng hủy."); return

                for en, (action, canonical) in dlg.result.items():
                    if action == "skip":
                        final_map[en] = None
                    elif action == "new":
                        unit = excel_units.get(en, "Viên")
                        mid = upsert_med(conn, canonical, unit)
                        if norm(canonical) != norm(en): add_alias(conn, mid, en)
                        final_map[en] = mid
                    elif action == "map":
                        mid = find_med(conn, canonical) or upsert_med(conn, canonical, excel_units.get(en, "Viên"))
                        add_alias(conn, mid, en)
                        final_map[en] = mid

            # Pass 3: nhập vào DB
            for fi, res in file_results:
                station = fi["station"]; period = fi["period"]
                src = Path(fi["filepath"]).name

                # Xóa dữ liệu cũ nếu file này đã từng nhập (tránh trùng lặp)
                existing_ids = [r[0] for r in conn.execute(
                    "SELECT id FROM encounters WHERE source_file=? AND station=?",
                    (src, station)
                ).fetchall()]
                if existing_ids:
                    ph = ",".join("?" * len(existing_ids))
                    conn.execute(
                        f"DELETE FROM prescriptions WHERE encounter_id IN ({ph})",
                        existing_ids)
                    conn.execute(
                        "DELETE FROM encounters WHERE source_file=? AND station=?",
                        (src, station))
                    conn.execute(
                        "DELETE FROM station_medicine_meta WHERE station=? AND period=?",
                        (station, period))
                    conn.commit()
                    self._log_msg(
                        f"   🔄 [{station}] Xóa {len(existing_ids)} ca cũ, nhập lại...")

                # Lưu medicine meta (tồn đầu, nhập) cho trạm này
                for med in res["medicines"]:
                    mid = find_med(conn, med["name"]) or final_map.get(med["name"])
                    if not mid: continue
                    # Cập nhật unit từ Excel (nguồn chính xác nhất)
                    u = str(med.get("unit", "") or "").strip()
                    if u:
                        conn.execute("UPDATE medicines SET unit=? WHERE id=?", (u, mid))
                    try:
                        conn.execute("""
                            INSERT OR REPLACE INTO station_medicine_meta
                                (station, period, medicine_id, expiry_date, beg_stock, received, issued)
                            VALUES (?,?,?,?,?,?,?)
                        """, (station, period, mid,
                              med.get("expiry",""), med.get("beg_stock",0),
                              med.get("received",0), med.get("issued",0)))
                    except Exception as me:
                        self._log_msg(f"   ⚠️ Meta {med['name'][:20]}: {me}")

                for enc in res["encounters"]:
                    conn.execute("INSERT INTO patients (id,name) VALUES (?,?) "
                                 "ON CONFLICT(id) DO UPDATE SET name=excluded.name",
                                 (enc["patient_id"], enc["patient_name"]))
                    cur = conn.execute("""
                        INSERT INTO encounters
                            (patient_id, patient_name, enc_date, time_in, time_out,
                             diagnosis, disease_group, status, station, source_file)
                        VALUES (?,?,?,?,?,?,?,?,?,?)
                    """, (enc["patient_id"], enc["patient_name"],
                          enc["date"].strftime("%Y-%m-%d"),
                          enc["time_in"], enc["time_out"],
                          enc["diagnosis"], enc["disease_group"],
                          enc["status"], station, src))
                    eid = cur.lastrowid

                    for med_name, qty in enc["medicines"].items():
                        mid = find_med(conn, med_name) or final_map.get(med_name)
                        if mid and qty > 0:
                            conn.execute(
                                "INSERT INTO prescriptions (encounter_id,medicine_id,quantity,station) VALUES (?,?,?,?)",
                                (eid, mid, qty, station))
                    total_enc += 1

                conn.commit()
                # Ghi nhận file đã nhập để tránh trùng lặp lần sau
                conn.execute("""
                    INSERT OR REPLACE INTO imported_files
                        (filename, station, period, imported_at)
                    VALUES (?,?,?,?)
                """, (src, station, period, datetime.now().isoformat()))
                conn.commit()
                self._log_msg(f"   ✅ [{station}] {len(res['encounters'])} ca nhập xong.")

            self._log_msg(f"🎉 Hoàn tất! {total_enc} ca khám đã vào database.")
            # Xóa danh sách file sau khi nhập thành công (tránh nhập lại vô tình)
            self._files.clear()
            self._flist.delete(0, "end")
            self._refresh_stats()
            messagebox.showinfo("Hoàn tất", f"Đã nhập {total_enc} ca khám.")

        except Exception as e:
            import traceback
            self._log_msg(f"❌ {e}\n{traceback.format_exc()}")
            messagebox.showerror("Lỗi", str(e))
        finally:
            conn.close()

    def _get_period_label(self) -> str:
        # Lấy tháng/kỳ từ file đầu tiên trong list (hoặc để trống)
        if self._files:
            return self._files[0]["period"]
        return ""

    def _export_patient(self):
        conn = get_db()
        count = conn.execute("SELECT COUNT(*) FROM encounters").fetchone()[0]
        conn.close()
        if count == 0:
            messagebox.showwarning("Chưa có dữ liệu", "Hãy nhập dữ liệu trước."); return
        p = filedialog.asksaveasfilename(title="Lưu danh sách bệnh nhân",
                                          defaultextension=".xlsx",
                                          filetypes=[("Excel","*.xlsx")],
                                          initialfile=f"BaoCao_BenhNhan_{datetime.now():%Y%m%d}.xlsx")
        if not p: return
        try:
            generate_patient_report(p, self._get_period_label())
            self._log_msg(f"✅ Danh sách bệnh nhân: {Path(p).name}")
            messagebox.showinfo("Xuất thành công", f"Đã xuất:\n{p}")
        except Exception as e:
            self._log_msg(f"❌ {e}"); messagebox.showerror("Lỗi", str(e))

    def _export_disease(self):
        conn = get_db()
        count = conn.execute("SELECT COUNT(*) FROM encounters").fetchone()[0]
        conn.close()
        if count == 0:
            messagebox.showwarning("Chưa có dữ liệu", "Hãy nhập dữ liệu trước."); return
        p = filedialog.asksaveasfilename(title="Lưu báo cáo mặt bệnh",
                                          defaultextension=".xlsx",
                                          filetypes=[("Excel","*.xlsx")],
                                          initialfile=f"BaoCao_MatBenh_{datetime.now():%Y%m%d}.xlsx")
        if not p: return
        try:
            generate_disease_report(p, self._get_period_label())
            self._log_msg(f"✅ Báo cáo mặt bệnh: {Path(p).name}")
            messagebox.showinfo("Xuất thành công", f"Đã xuất:\n{p}")
        except Exception as e:
            self._log_msg(f"❌ {e}"); messagebox.showerror("Lỗi", str(e))

    def _export_medicine(self):
        conn = get_db()
        count = conn.execute("SELECT COUNT(*) FROM encounters").fetchone()[0]
        conn.close()
        if count == 0:
            messagebox.showwarning("Chưa có dữ liệu", "Hãy nhập dữ liệu trước."); return
        p = filedialog.asksaveasfilename(title="Lưu báo cáo thuốc",
                                          defaultextension=".xlsx",
                                          filetypes=[("Excel","*.xlsx")],
                                          initialfile=f"BaoCao_SuDungThuoc_{datetime.now():%Y%m%d}.xlsx")
        if not p: return
        try:
            generate_medicine_report(p, self._get_period_label())
            self._log_msg(f"✅ Báo cáo thuốc: {Path(p).name}")
            messagebox.showinfo("Xuất thành công", f"Đã xuất:\n{p}")
        except Exception as e:
            self._log_msg(f"❌ {e}"); messagebox.showerror("Lỗi", str(e))

    def _reset(self):
        conn = get_db(); s = db_stats(conn); conn.close()
        if not messagebox.askyesno("Xóa toàn bộ",
            f"Xóa {s['encounters']} ca khám, {s['medicines']} thuốc?\nKHÔNG THỂ hoàn tác!"): return
        if DB_PATH.exists(): DB_PATH.unlink()
        get_db().close()
        self._log_msg("🗑 DB đã reset."); self._refresh_stats()

    def _disease_group_dlg(self):
        """Mở dialog chuẩn hóa nhóm bệnh."""
        conn = get_db()
        cnt = conn.execute("SELECT COUNT(DISTINCT disease_group) FROM encounters WHERE disease_group!=''").fetchone()[0]
        conn.close()
        if cnt == 0:
            messagebox.showinfo("Chưa có dữ liệu", "Hãy nhập dữ liệu trước để xem nhóm bệnh."); return

        dlg = DiseaseGroupDialog(self)
        self.wait_window(dlg)
        if not dlg.confirmed or not dlg.result: return

        # Áp dụng mapping: UPDATE encounters SET disease_group=? WHERE disease_group=?
        changes = {src: dst for src, dst in dlg.result.items() if src != dst}
        if not changes:
            self._log_msg("ℹ️ Không có thay đổi nhóm bệnh."); return
        conn = get_db()
        try:
            for src, dst in changes.items():
                conn.execute("UPDATE encounters SET disease_group=? WHERE disease_group=?", (dst, src))
                self._log_msg(f"  🏷 '{src}' → '{dst}'")
            conn.commit()
            self._log_msg(f"✅ Đã cập nhật {len(changes)} nhóm bệnh.")
        finally:
            conn.close()
        self._refresh_stats()


if __name__ == "__main__":
    App().mainloop()
